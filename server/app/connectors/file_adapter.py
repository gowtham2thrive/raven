"""
File Adapter — CSV, Excel, and PDF evidence ingestion.

Handles three file formats:
- CSV: Configurable delimiter, encoding, header row
- Excel (.xlsx): Sheet selection, header row
- PDF: Text extraction via pdfplumber, structured table extraction

All formats produce the same output: a list of dicts that the
field mapping engine translates to canonical evidence fields.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from app.connectors.adapter_registry import register_adapter
from app.connectors.base_adapter import BaseAdapter
from app.core.integration_schemas import (
    FileUploadConfig,
    IntegrationTestResult,
)
from app.core.integration_types import (
    IntegrationFileError,
    IntegrationType,
)

logger = logging.getLogger(__name__)

MAX_SAMPLE_RECORDS = 5
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB


# ═══════════════════════════════════════════════════════════════
#  CSV ADAPTER
# ═══════════════════════════════════════════════════════════════


@register_adapter(IntegrationType.CSV_FILE)
class CsvFileAdapter(BaseAdapter):
    """Reads evidence data from uploaded CSV files."""

    def __init__(self, config: dict[str, Any], integration_id: str = ""):
        self._integration_id = integration_id
        self._config = FileUploadConfig.model_validate(config)
        self._discovered_fields: list[str] = []
        self._cached_records: list[dict[str, Any]] | None = None

    def test_connection(self) -> IntegrationTestResult:
        """Validate the CSV file and return sample data."""
        def _test():
            self._validate_file()
            records = self._parse_csv()
            self._cached_records = records
            sample = records[:MAX_SAMPLE_RECORDS]
            fields = self._discover_fields_from_records(sample)
            self._discovered_fields = fields

            return IntegrationTestResult(
                success=True,
                message=f"CSV parsed successfully. Found {len(records)} row(s), {len(fields)} field(s).",
                sample_data=sample,
                discovered_fields=fields,
                record_count=len(records),
            )

        return self._timed_test(_test)

    def fetch_raw_data(self, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        """Return all records from the CSV file."""
        if self._cached_records is not None:
            return self._cached_records

        self._validate_file()
        records = self._parse_csv()
        self._cached_records = records
        return records

    def get_sample_fields(self) -> list[str]:
        """Return field names from the CSV header."""
        if self._discovered_fields:
            return self._discovered_fields

        try:
            result = self.test_connection()
            return result.discovered_fields
        except Exception:
            return []

    def _validate_file(self) -> None:
        """Validate the file exists and is within size limits."""
        file_path = Path(self._config.stored_path)
        if not file_path.exists():
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File not found at {self._config.stored_path}",
            )

        size = file_path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File too large: {size / 1024 / 1024:.1f} MB (max {MAX_FILE_SIZE_BYTES / 1024 / 1024:.0f} MB)",
            )

    def _parse_csv(self) -> list[dict[str, Any]]:
        """Parse the CSV file into a list of dicts."""
        file_path = Path(self._config.stored_path)

        try:
            with open(file_path, encoding=self._config.encoding, newline="") as f:
                if self._config.has_header:
                    reader = csv.DictReader(f, delimiter=self._config.delimiter)
                    records = []
                    for row in reader:
                        cleaned_row = {}
                        for k, v in row.items():
                            key_str = str(k) if k is not None else "extra"
                            cleaned_row[key_str] = v
                        records.append(cleaned_row)
                    return records
                else:
                    reader = csv.reader(f, delimiter=self._config.delimiter)
                    rows = list(reader)
                    if not rows:
                        return []
                    # Generate column names: col_0, col_1, ...
                    headers = [f"col_{i}" for i in range(len(rows[0]))]
                    return [dict(zip(headers, row)) for row in rows]

        except UnicodeDecodeError as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"Encoding error (expected {self._config.encoding}): {e}",
            ) from e
        except csv.Error as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"CSV parse error: {e}",
            ) from e

    @staticmethod
    def _discover_fields_from_records(records: list[dict]) -> list[str]:
        """Collect all unique field names across records."""
        fields: set[str] = set()
        for record in records:
            fields.update(record.keys())
        return sorted(fields)


# ═══════════════════════════════════════════════════════════════
#  EXCEL ADAPTER
# ═══════════════════════════════════════════════════════════════


@register_adapter(IntegrationType.EXCEL_FILE)
class ExcelFileAdapter(BaseAdapter):
    """Reads evidence data from uploaded Excel (.xlsx) files."""

    def __init__(self, config: dict[str, Any], integration_id: str = ""):
        self._integration_id = integration_id
        self._config = FileUploadConfig.model_validate(config)
        self._discovered_fields: list[str] = []
        self._cached_records: list[dict[str, Any]] | None = None

    def test_connection(self) -> IntegrationTestResult:
        """Validate the Excel file and return sample data."""
        def _test():
            self._validate_file()
            records = self._parse_excel()
            self._cached_records = records
            sample = records[:MAX_SAMPLE_RECORDS]
            fields = self._discover_fields_from_records(sample)
            self._discovered_fields = fields

            return IntegrationTestResult(
                success=True,
                message=f"Excel parsed successfully. Found {len(records)} row(s), {len(fields)} field(s).",
                sample_data=sample,
                discovered_fields=fields,
                record_count=len(records),
            )

        return self._timed_test(_test)

    def fetch_raw_data(self, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        """Return all records from the Excel file."""
        if self._cached_records is not None:
            return self._cached_records

        self._validate_file()
        records = self._parse_excel()
        self._cached_records = records
        return records

    def get_sample_fields(self) -> list[str]:
        """Return field names from the Excel header."""
        if self._discovered_fields:
            return self._discovered_fields

        try:
            result = self.test_connection()
            return result.discovered_fields
        except Exception:
            return []

    def _validate_file(self) -> None:
        """Validate the file exists and is within size limits."""
        file_path = Path(self._config.stored_path)
        if not file_path.exists():
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File not found at {self._config.stored_path}",
            )

        size = file_path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File too large: {size / 1024 / 1024:.1f} MB (max {MAX_FILE_SIZE_BYTES / 1024 / 1024:.0f} MB)",
            )

    def _parse_excel(self) -> list[dict[str, Any]]:
        """Parse the Excel file into a list of dicts."""
        try:
            import openpyxl
        except ImportError as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                "openpyxl is required for Excel files. Install with: pip install openpyxl",
            ) from e

        file_path = Path(self._config.stored_path)

        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)

            if self._config.sheet_name:
                if self._config.sheet_name not in wb.sheetnames:
                    raise IntegrationFileError(
                        self._integration_id,
                        self._config.filename,
                        f"Sheet '{self._config.sheet_name}' not found. Available: {wb.sheetnames}",
                    )
                ws = wb[self._config.sheet_name]
            else:
                ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not rows:
                return []

            header_idx = self._config.header_row
            if header_idx >= len(rows):
                raise IntegrationFileError(
                    self._integration_id,
                    self._config.filename,
                    f"Header row {header_idx} exceeds total rows {len(rows)}",
                )

            # Ensure unique and non-empty header names
            raw_headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[header_idx])]
            seen_headers: dict[str, int] = {}
            headers = []
            for h in raw_headers:
                if not h:
                    h = f"col_{len(headers)}"
                if h in seen_headers:
                    seen_headers[h] += 1
                    headers.append(f"{h}_{seen_headers[h]}")
                else:
                    seen_headers[h] = 0
                    headers.append(h)

            records = []
            for row in rows[header_idx + 1:]:
                if not any(v is not None and str(v).strip() != "" for v in row):
                    continue
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        # Convert datetime objects to ISO strings for consistency
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        record[headers[i]] = value
                records.append(record)

            return records

        except IntegrationFileError:
            raise
        except Exception as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"Excel parse error: {e}",
            ) from e

    @staticmethod
    def _discover_fields_from_records(records: list[dict]) -> list[str]:
        """Collect all unique field names across records."""
        fields: set[str] = set()
        for record in records:
            fields.update(record.keys())
        return sorted(fields)


# ═══════════════════════════════════════════════════════════════
#  PDF ADAPTER
# ═══════════════════════════════════════════════════════════════


@register_adapter(IntegrationType.PDF_FILE)
class PdfFileAdapter(BaseAdapter):
    """Extracts evidence data from uploaded PDF files.

    Uses pdfplumber for text and table extraction.
    Returns either extracted tables (as list of dicts) or
    raw text pages (as list of {'page': int, 'text': str}).
    """

    def __init__(self, config: dict[str, Any], integration_id: str = ""):
        self._integration_id = integration_id
        self._config = FileUploadConfig.model_validate(config)
        self._discovered_fields: list[str] = []
        self._cached_records: list[dict[str, Any]] | None = None

    def test_connection(self) -> IntegrationTestResult:
        """Validate the PDF file and return extracted data preview."""
        def _test():
            self._validate_file()
            records = self._extract_pdf()
            self._cached_records = records
            sample = records[:MAX_SAMPLE_RECORDS]
            fields = self._discover_fields_from_records(sample)
            self._discovered_fields = fields

            return IntegrationTestResult(
                success=True,
                message=f"PDF processed successfully. Extracted {len(records)} record(s).",
                sample_data=sample,
                discovered_fields=fields,
                record_count=len(records),
            )

        return self._timed_test(_test)

    def fetch_raw_data(self, params: dict[str, Any] | None = None) -> list[dict[str, Any]]:
        """Return all extracted records from the PDF."""
        if self._cached_records is not None:
            return self._cached_records

        self._validate_file()
        records = self._extract_pdf()
        self._cached_records = records
        return records

    def get_sample_fields(self) -> list[str]:
        """Return field names discovered from the PDF."""
        if self._discovered_fields:
            return self._discovered_fields

        try:
            result = self.test_connection()
            return result.discovered_fields
        except Exception:
            return []

    def _validate_file(self) -> None:
        """Validate the file exists and is within size limits."""
        file_path = Path(self._config.stored_path)
        if not file_path.exists():
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File not found at {self._config.stored_path}",
            )

        size = file_path.stat().st_size
        if size > MAX_FILE_SIZE_BYTES:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"File too large: {size / 1024 / 1024:.1f} MB (max {MAX_FILE_SIZE_BYTES / 1024 / 1024:.0f} MB)",
            )

    def _extract_pdf(self) -> list[dict[str, Any]]:
        """Extract tables and text from the PDF.

        Strategy:
        1. Try to extract structured tables first
        2. If no tables found, fall back to page-by-page text extraction
        """
        try:
            import pdfplumber
        except ImportError as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                "pdfplumber is required for PDF files. Install with: pip install pdfplumber",
            ) from e

        file_path = Path(self._config.stored_path)
        records: list[dict[str, Any]] = []

        try:
            with pdfplumber.open(str(file_path)) as pdf:
                # Attempt table extraction if configured
                if self._config.extract_tables:
                    table_records = self._extract_tables(pdf)
                    if table_records:
                        return table_records

                # Fall back to text extraction
                for page_num, page in enumerate(pdf.pages, start=1):
                    text = page.extract_text()
                    if text and text.strip():
                        records.append({
                            "page_number": page_num,
                            "text": text.strip(),
                            "char_count": len(text.strip()),
                        })

        except IntegrationFileError:
            raise
        except Exception as e:
            raise IntegrationFileError(
                self._integration_id,
                self._config.filename,
                f"PDF extraction error: {e}",
            ) from e

        return records

    @staticmethod
    def _extract_tables(pdf) -> list[dict[str, Any]]:
        """Extract structured tables from PDF pages.

        Returns a list of dicts if any tables with headers are found.
        """
        all_records: list[dict[str, Any]] = []

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table or len(table) < 2:
                    continue

                # Ensure unique and non-empty header names
                raw_headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(table[0])]
                seen_headers: dict[str, int] = {}
                headers = []
                for h in raw_headers:
                    if not h:
                        h = f"col_{len(headers)}"
                    if h in seen_headers:
                        seen_headers[h] += 1
                        headers.append(f"{h}_{seen_headers[h]}")
                    else:
                        seen_headers[h] = 0
                        headers.append(h)

                for row in table[1:]:
                    if row and any(cell for cell in row):
                        record = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                record[headers[i]] = str(cell).strip() if cell else ""
                        all_records.append(record)

        return all_records

    @staticmethod
    def _discover_fields_from_records(records: list[dict]) -> list[str]:
        """Collect all unique field names across records."""
        fields: set[str] = set()
        for record in records:
            fields.update(record.keys())
        return sorted(fields)
